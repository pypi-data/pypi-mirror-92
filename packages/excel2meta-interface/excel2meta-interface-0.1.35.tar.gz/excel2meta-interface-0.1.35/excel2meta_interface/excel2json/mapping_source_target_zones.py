from openpyxl import load_workbook
from datetime import datetime
from resources import excel_headers
import sys
from excel2meta_interface.utils import excel


class MappingSource2TargetZone:
    start_time = datetime.now()

    def __init__(self, excel_file, sheet):
        self.excel_file = excel_file
        self.excel_sheet = sheet
        self.work_book = None
        self.work_sheet = None
        # self.excel2ref = excel2references.Excel2Reference()
        self.excel_utils = excel.Excel()

    def get_source2target_mapping(self):
        result, self.work_book, self.work_sheet = self.excel_utils.read_excel(file=self.excel_file, sheet=self.excel_sheet)
        xls_ref_index = self.excel_utils.search_value_in_row_index(work_sheet=self.work_sheet
                                                                   ,search_string=excel_headers.mapping_overview_consolidated["xls_file"])[0]

        if result["code"] != "OK":
            return None
        excel_list = []
        for i in range(2, self.work_sheet.max_row + 1):
            d = {}
            for j in range(1, self.work_sheet.max_column + 1):
                cell_value_class = self.work_sheet.cell(1,j).value
                cell_value_id = self.work_sheet.cell(i,j).value
                d[cell_value_class] = cell_value_id
            excel_list.append(d)

        return excel_list


if __name__ == "__main__":
    excel_file = sys.argv[1]
    sheet = sys.argv[2]
    mapping_ref = MappingSource2TargetZone(excel_file=excel_file, sheet=sheet)
    the_list = mapping_ref.get_source2target_mapping()
    print(the_list)