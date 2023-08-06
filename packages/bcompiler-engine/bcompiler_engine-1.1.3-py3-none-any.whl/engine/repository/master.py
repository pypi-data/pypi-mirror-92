# repository/master.py
import logging
from pathlib import Path

from openpyxl import Workbook

from engine.config import Config

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s: %(levelname)s - %(message)s",
    datefmt="%d-%b-%y %H:%M:%S",
)
logger = logging.getLogger(__name__)


class ValidationOnlyRepository:
    def __init__(self, data, output_file_name=None):
        self.data = data

    # Does nothing
    def save(self) -> None:
        logger.info("No output file produced as not requested.")


class MasterOutputRepository:
    def __init__(self, data, output_file_name):
        self.data = data
        self.output_filename = output_file_name

    def save(self) -> None:
        _master_return_reference = Config.config_parser["DEFAULT"][
            "return reference name"
        ]
        output_path = Path(Config.PLATFORM_DOCS_DIR) / "output"
        wb = Workbook()
        ws = wb.active
        ws.title = "Master"
        ws["A1"].value = _master_return_reference
        _keys = [x[0] for x in list(self.data[0].values())[0]]
        # col A
        for i, k in enumerate(_keys, start=2):
            ws.cell(column=1, row=i, value=k)
        # other cols
        for counter, file_data in enumerate(self.data, start=2):
            ws.cell(
                column=counter, row=1, value=list(file_data.keys())[0].split(".")[0]
            )
            _key_value_lst = list(file_data.values())[0]
            for idx, tup in enumerate(_key_value_lst, start=2):
                ws.cell(column=counter, row=idx, value=tup[1])
        wb.save(output_path / self.output_filename)
        logger.info(
            "{} successfully created in {}\n".format(self.output_filename, output_path)
        )
