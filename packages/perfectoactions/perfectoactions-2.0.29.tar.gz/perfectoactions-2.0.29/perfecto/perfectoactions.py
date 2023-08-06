#!/usr/bin/env python
import argparse
import base64
import datetime
import io
import json
import multiprocessing
import os
import platform
import re
import shutil
import ssl
import sys
import tempfile
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
import uuid
import webbrowser
import xml.etree.ElementTree as ETree
from datetime import timedelta
from multiprocessing import Pool, Process, freeze_support
from xml.dom import minidom

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas
import pylab as pl
import requests
import tzlocal
from colorama import init
from datetimerange import DateTimeRange
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from pandas.plotting import table
from termcolor import colored

""" Microsoft Visual C++ required, cython required for pandas installation, """
TEMP_DIR = "/tmp" if platform.system() == "Darwin" else tempfile.gettempdir()
# Do not change these variable
RESOURCE_TYPE = "handsets"
RESOURCE_TYPE_USERS = "users"
RESOURCE_TYPE_CRADLES = "cradles"
REPOSITORY_RESOURCE_TYPE = "repositories/media"
RESOURCE_TYPE_RESERVATIONS = "reservations"


def send_request(url):
    """send request"""
    #     print("Submitting", url)
    device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
    if (
        "All devices" in device_list_parameters
        or "Available devices only" in device_list_parameters
    ):
        response = urllib.request.urlopen(url)
    else:
        response = urllib.request.urlopen(url.replace(" ", "%20"))
    #    rc = response.getcode()
    #    print("rc =", rc)
    return response


def send_request_with_json_response(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    maps = json.loads(text)
    print("Response:\n" + str(maps))
    return maps


def as_text(value):
    """as texts"""
    if value is None:
        return ""
    return str(value)


def convertxmlToXls(xml, dict_keys, filename):
    """
            Checks if file exists, parses the file and extracts the needed data
            returns a 2 dimensional list without "header"
    """
    root = ETree.fromstring(xml)
    headers = []
    finalHeaders = []
    if dict_keys is None:
        for child in root:
            headers.append({x.tag for x in root.findall(child.tag + "/*") })
    else:
        headers = dict_keys
    headers = headers[0]
    mdlist = []
    for child in root:
        temp = []
        for key in sorted(headers):
            try:
                finalHeaders.append(key)
                temp.append(child.find(key).text)
            except Exception as e:
                temp.append("-")
        mdlist.append(temp)
    """
    Generates excel file with given data
    mdlist: 2 Dimensional list containing data
    """
    wb = Workbook()
    ws = wb.active
    for i, row in enumerate(mdlist):
        for j, value in enumerate(row):
            ws.cell(row=i + 1, column=j + 1).value = value
    ws.insert_rows(0)
    # generates header
    i = 0
    finalHeaders = list(dict.fromkeys(finalHeaders))
    for i, value in enumerate(finalHeaders):
        ws.cell(1, column=i + 1).value = value
        ws.cell(1, column=i + 1).alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    newfilename = os.path.abspath(filename)
    wb.save(newfilename)
    return


def user_condition(df):
    cols = list(df)
    if len(cols) > 0:
        cols = [cols[-1]] + cols[:-1]
        df = df[cols]
        df = df.replace(np.nan, "", regex=True)
        df = df[~df["email"].str.contains("perfectomobile.com")]
        df = df.sort_values(by="firstName")
    return df


def convertjsonToXls(json_text, dict_keys, filename):
    jsonfile = "user_results.json"
    file = os.path.join(TEMP_DIR, "results", jsonfile)
    f = open(file, "w+")
    f.write(str(json_text))
    f.close()
    data = json.load(open(file))
    sys.stdout.flush()
    pandas.set_option("display.max_columns", 6)
    # pandas.set_option('display.max_colwidth', 120)
    pandas.set_option("colheader_justify", "left")
    df = pandas.DataFrame(data["users"])
    if len(df.index) < 1:
        raise Exception(
            "There are no users who match the expected conditions "
            + os.environ["USER_LIST_PARAMETERS"]
        )
        sys.exit(-1)
    df.drop(
        [
            "username",
            "authentication",
            "gender",
            "phoneNumberExt",
            "location",
            "stateCode",
            "state",
        ],
        axis=1,
        inplace=True,
        errors="ignore",
    )
    df = user_condition(df)
    df.to_excel(filename, index=False)
    wb = Workbook()
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    newfilename = os.path.abspath(filename)
    wb.save(newfilename)
    return df


def send_request_with_xml_response(url):
    """send request"""
    print("Attempting url: \n" + str(url))
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    xmldoc = minidom.parseString(decoded)
    return xmldoc


def send_request_to_xlsx(url, filename):
    """send_request_to_xlsx"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    if any(["=list" in url, "=users" in url]):
        filename = os.path.join(TEMP_DIR, "output", filename)
        convertxmlToXls(decoded, None, filename)

def send_cradles_request_to_xlsx(url, command, filename):
    """send_cradles_request_to_xlsx"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    xmldoc = minidom.parseString(decoded)
    cradleElements = xmldoc.getElementsByTagName("cradle")
    nCradles = len(cradleElements) 
    print ("found " + str(nCradles) + " cradles")
    if (nCradles >= 1):
        idElements = xmldoc.getElementsByTagName("id")
        resource = idElements[0].firstChild.data
        url = get_url(RESOURCE_TYPE_CRADLES, resource, command)
        response = urllib.request.urlopen(url)
        text = response.read().decode("utf-8")
        xmldoc = minidom.parseString(text)
        locationElements = xmldoc.getElementsByTagName("location")
        location = locationElements[0].firstChild.data
        print ("the location of cradle " + resource + " is " + location)
        filename = os.path.join(TEMP_DIR, "output", filename)
        print(text)
        convertxmlToXls(text, None, filename)

def send_jsonrequest_to_xlsx(url, filename):
    """send_request_to_xlsx"""
    print("Attempting User list API: \n" + str(url))
    try:
        response = send_request(url)
        print("response: " + str(response))
        decoded = response.read().decode("utf-8")
        if any(["=list" in url, "=users" in url]):
            filename = os.path.join(TEMP_DIR, "output", filename)
        return convertjsonToXls(decoded, None, filename)
    except Exception as e:
        return e
    


def send_request2(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    return text


def get_url(resource, resource_id, operation):
    """get url """
    cloudname = os.environ["CLOUDNAME"]
    url = "https://" + cloudname + ".perfectomobile.com/services/" + resource
    if resource_id != "":
        url += "/" + str(resource_id)
    token = os.environ["TOKEN"]
    if "eyJhb" in token:
        query = urllib.parse.urlencode({"operation": operation, "securityToken": token})
    else:
        if ":" not in token:
            raise Exception(
                "Please pass your perfecto credentials in the format user:password as -s parameter value. Avoid using special characters such as :,@. in passwords!"
            )
            sys.exit(-1)
        else:
            user = token.split(":")[0]
            pwd = token.split(":")[1]
            query = urllib.parse.urlencode(
                {"operation": operation, "user": user, "password": pwd}
            )
    url += "?" + query
    return url


def flatten_json(nested_json, exclude=[""]):
    """Flatten json object with nested keys into a single level.
        Args:
            nested_json: A nested json object.
            exclude: Keys to exclude from output.
        Returns:
            The flattened json object if successful, None otherwise.
    """
    out = {}

    def flatten(x, name="", exclude=exclude):
        if type(x) is dict:
            for a in x:
                if a not in exclude:
                    flatten(x[a], name + a + "/")
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + "/")
                i += 1
        else:
            out[name[:-1]] = x
    flatten(nested_json)
    return out

def getregex_output(response, pattern1, pattern2):
    """regex"""
    print(str(response))
    matches = re.finditer(pattern1, response, re.MULTILINE)
    for match in matches:
        match_item = str(re.findall(pattern2, match.group()))
        match_item = match_item.replace(':"', "").replace('"', "")
        match_item = match_item.replace("'", "").replace("[", "")
        match_item = (
            match_item.replace("]", "")
            .replace(",test", "")
            .replace(",timer.system", "")
            .replace('description":"', "")
        )
        return str(match_item)

def getTimeofExecution(response, pattern1, pattern2):
    """getTimestamp"""
    matches = re.finditer(pattern1, response, re.MULTILINE)
    for match in matches:
        match_item = str(re.findall(pattern2, match.group()))
        match_item = match_item.replace('timestamp[0]=', "").replace("'","").replace("[","").replace("]","").split(",")[0]
        return str(time.strftime("%d %b %Y %H:%M:%S", 
                    time.localtime(int(match_item)/ 1000)))

def device_command(exec_id, device_id, operation):
    """Runs device command"""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + "device"
    url += "&subcommand=" + operation
    url += "&param.deviceId=" + device_id
    send_request_with_json_response(url)


def end_execution(exec_id):
    """End execution"""
    url = get_url("executions/" + str(exec_id), "", "end")
    send_request_with_json_response(url)


def start_exec():
    """start execution"""
    url = get_url("executions", "", "start")
    response = send_request2(url)
    exec_id = getregex_output(response, r"executionId\"\:\"[\w\d@.-]+\"", ':".*$')
    return exec_id


def get_device_list_response(resource, command, status, in_use):
    """get_device_list_response"""
    url = get_url(resource, "", command)
    url += "&status=" + status
    if in_use != "":
        url += "&inUse=" + in_use
    if len(os.environ["DEVICE_LIST_PARAMETERS"].split(":")) >= 2:
        for item in os.environ["DEVICE_LIST_PARAMETERS"].split(";"):
            if ":" in item:
                url += "&" + item.split(":")[0] + "=" + item.split(":")[1]
    xmldoc = send_request_with_xml_response(url)
    return xmldoc

def get_reservations_list_response(resource, command, end, admin):
    """get_reservations_list_response"""
    url = get_url(resource, "", command)
    # url += "&responseFormat=xml"
    url += "&startTime=" + str(int((time.mktime(datetime.datetime.now().timetuple()))) *  1000)
    if end != "":
        endTime = datetime.datetime.fromtimestamp(int(time.mktime(datetime.datetime.now().timetuple()))) + datetime.timedelta(days=int(end))
        url += "&endTime=" + str(int(endTime.timestamp()) * 1000)
    if admin != "":
        url += "&admin=" + admin
    print("Attempting get reservations list API: \n" + url)
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    return decoded

def get_xml_to_xlsx(resource, command, filename):
    """get_xml_to_xlsx"""
    url = get_url(resource, "", command)
    send_request_to_xlsx(url, filename)
    sys.stdout.flush()

def get_cradles_xml_to_xlsx(resource, command, filename):
    """get_cradles_xml_to_xlsx"""
    url = get_url(resource, "", "list")
    # url = get_url(resource, "", command)
    send_cradles_request_to_xlsx(url, command, filename)
    sys.stdout.flush()


def get_json_to_xlsx(resource, command, filename):
    """get_json_to_xlsx"""
    url = get_url(resource, "", command)
    if "All users" not in os.environ["USER_LIST_PARAMETERS"]:
        if len(os.environ["USER_LIST_PARAMETERS"].split(":")) >= 2:
            for item in os.environ["USER_LIST_PARAMETERS"].split(";"):
                if ":" in item:
                    url += "&" + item.split(":")[0] + "=" + item.split(":")[1]
    return send_jsonrequest_to_xlsx(url.replace(" ", "%20"), filename)


def get_device_ids(xmldoc):
    """get_device_ids"""
    device_ids = xmldoc.getElementsByTagName("deviceId")
    return device_ids


def get_handset_count(xmldoc):
    """get_handset_count"""
    handset_elements = xmldoc.getElementsByTagName("handset")
    return len(handset_elements)


def exec_command(exec_id, device_id, cmd, subcmd):
    """exec_commands"""
    status = ""
    timeofExec = ""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + cmd
    url += "&subcommand=" + subcmd
    url += "&param.deviceId=" + device_id
    response = send_request2(url)
    status = getregex_output(
        response,
        r"(description\"\:\".*\",\"timer.system|returnValue\"\:\".*\",\"test)",
        ':".*$',
    )
    if("Failed" not in status):
        timeofExec = getTimeofExecution(
            response,
            r"reportPdfUrl\"\:[\s]?\".*\",[\s]?\"executionId",
            'timestamp\[0\]\=[\d]+',
        )
    return str(status).replace(",", " "), str(timeofExec)

def create_reservation(resource, resource_id, operation):
    """create_reservation"""
    cloudname = os.environ["CLOUDNAME"]
    url = "https://" + cloudname + ".perfectomobile.com/services/" + resource
    token = os.environ["TOKEN"]
    today = datetime.datetime.now() + timedelta(seconds = 20)
    startTime = int((time.mktime(today.timetuple())))
    reserve_time = int(str(os.environ["RESERVE_TIME"]))
    endTime = int((datetime.datetime.fromtimestamp(startTime) + timedelta(minutes = reserve_time)).timestamp())
    if "eyJhb" in token:
        query = urllib.parse.urlencode({"operation": operation, "securityToken": token, "resourceIds": resource_id, "startTime" : startTime*1000, "endTime": endTime*1000})
    else:
        if ":" in token: 
            user = token.split(":")[0]
            pwd = token.split(":")[1]
            query = urllib.parse.urlencode                                                                                                                                                                             
        else:
            raise Exception(
                "Please pass your perfecto credentials in the format user:password as -s parameter value. Avoid using special characters such as :,@. in passwords!"
            )
    url += "?" + query
    print(url)
    return url

def perform_actions(deviceid_color):
    """perform_actions"""
    get_network_settings = os.environ["GET_NETWORK_SETTINGS"]
    deviceid_color = str(deviceid_color)
    device_id = deviceid_color.split("||", 1)[0]
    color = deviceid_color.split("||", 1)[1]
    desc = deviceid_color.split("||", 2)[2]
    fileName = device_id + ".txt"
    file = os.path.join(TEMP_DIR, "results", fileName)
    timeofExec = ""
    try:
        status = "Results="
        # update dictionary
        url = get_url(RESOURCE_TYPE, device_id, "info")
        xmldoc = send_request_with_xml_response(url)
  
        modelElements = xmldoc.getElementsByTagName("model")
        manufacturerElements = xmldoc.getElementsByTagName("manufacturer")
        model = modelElements[0].firstChild.data
        manufacturer = manufacturerElements[0].firstChild.data
        osElements = xmldoc.getElementsByTagName("os")
        osDevice = osElements[0].firstChild.data
        try:
            descriptionElements = xmldoc.getElementsByTagName("description")
            description = descriptionElements[0].firstChild.data
        except:
            description = ""
        try:
            osVElements = xmldoc.getElementsByTagName("osVersion")
            osVersion = osVElements[0].firstChild.data
        except:
            osVersion = "NA"
        osVersion = osDevice + " " + osVersion
        try:
            operatorElements = xmldoc.getElementsByTagName("operator")
            operator = operatorElements[0].childNodes[0].data
        except:
            operator = "NA"
        try:
            phElements = xmldoc.getElementsByTagName("phoneNumber")
            phoneNumber = phElements[0].firstChild.data
        except:
            phoneNumber = "NA"
        if "green" in color:
            reserve = os.environ["RESERVE"]
            if "True" in reserve:
                print(
                "Reserving device: "
                + device_id
                )
                try:
                    response = send_request(create_reservation(RESOURCE_TYPE_RESERVATIONS, device_id, "create"))
                    decoded = response.read().decode("utf-8")
                    print(decoded)
                    desc = "Reserved"
                    status += "RS:OK;"
                except Exception as e:
                    print("Reservation API error: " + str(e))
                    if("HTTP Error 500: Internal Server Error" in str(e).strip()):
                        status += "RS:"+"Reservations may exists;"
                    else:
                        status += "RS:"+str(e)+"!;"
                    final_string = (
                            "status=Reservation failed"
                            + ", deviceId='"
                            + device_id
                            + "', Manufacturer="
                            + str(manufacturer)
                            + "', model="
                            + str(model)
                            + ", version="
                            + str(osVersion)
                            + ", description="
                            + str(description)
                            + ", operator="
                            + str(operator)
                            + ", phoneNumber="
                            + str(phoneNumber)
                            + ",,,, "
                            + str(status)
                            + ", lastExecAt="
                            + str(timeofExec)
                        )
                    final_string = re.sub(r"^'|'$", "", final_string)
                    print(final_string)
                    f = open(file, "w+")
                    f.write(str(final_string))
                    f.close()
                    return final_string
        start_execution = os.environ["START_EXECUTION"]
        if "true" in start_execution.lower():
            # Get execution id
            EXEC_ID = start_exec()
            # open device:
            print("opening: " + model + ", device id: " + device_id)
            device_command(EXEC_ID, device_id, "open")
            cleanup = os.environ["CLEANUP"]
            temp = ""
            if "True" in cleanup:
                if not "iOS" in osDevice:
                    print("cleaning up: " + model + ", device id: " + device_id)
                    try:
                        output, temp =  exec_command(EXEC_ID, device_id, "device", "clean")
                        status += "clean:" + output
                    except Exception as e:
                        print(e)
                        status += "clean:Failed!"
                    status += ";"
                else:
                    status += "clean:NA;"
            reboot = os.environ["REBOOT"]
            if "True" in reboot:
                if all(
                    [
                        "Huawei" not in manufacturer,
                        "Xiaomi" not in manufacturer,
                        "Oppo" not in manufacturer,
                        "Motorola" not in manufacturer,
                        "OnePlus" not in manufacturer,
                    ]
                ):
                    print("rebooting: " + model + ", device id: " + device_id)
                    try:
                        output, timeofExec = exec_command(EXEC_ID, device_id, "device", "reboot")
                        status += "reboot:" +  output
                    except:
                        status += "reboot:Failed!"
                    status += ";"
                else:
                    print(model + " not applicable for rebooting")
                    status += "reboot:NA;"
            if "True" in get_network_settings:
                print(
                    "getting network status of : "
                    + model
                    + ", device id: "
                    + device_id
                )
                networkstatus = "airplanemode=Failed, wifi=Failed, data=Failed"
                try:
                    tempstatus, temp = exec_command(EXEC_ID, device_id, "network.settings", "get")
                    tempstatus = tempstatus.replace("{", "").replace("}", "")
                    if tempstatus.count(",") == 2 or tempstatus.count("  ") == 2:
                        if(tempstatus.count("  ") == 2):
                            networkstatus = ', e'.join(str(tempstatus).split("  "))
                        else:
                            networkstatus = tempstatus
                        status += "NW:OK"
                    else:
                        status += "NW:Failed!"
                except:
                    status += "NW:Failed!"
                status += ";"
            # Close device
            print("closing: " + model + ", device id: " + device_id)
            device_command(EXEC_ID, device_id, "close")
            # End execution
            end_execution(EXEC_ID)
            print(" -----------")
            print(timeofExec)
        else:
            networkstatus = ",,"

        if "True" in get_network_settings:
            final_string = (
                "status="
                + desc
                + ", deviceId='"
                + device_id
                + "', Manufacturer="
                + str(manufacturer)
                + "', model="
                + str(model)
                + ", version="
                + str(osVersion)
                + ", description="
                + str(description)
                + ", operator="
                + str(operator)
                + ", phoneNumber="
                + str(phoneNumber)
                + ", "
                + str(networkstatus)
                + ", "
                + str(status)
                + ", lastExecAt="
                + str(timeofExec)
            )
        else:
            final_string = (
                "status="
                + desc
                + ", deviceId='"
                + device_id
                + "', Manufacturer="
                + str(manufacturer)
                + "', model="
                + str(model)
                + ", version="
                + str(osVersion)
                + ", description="
                + str(description)
                + ", operator="
                + str(operator)
                + ", phoneNumber="
                + str(phoneNumber)
                + ",,,, "
                + str(status)
                + ", lastExecAt="
                + str(timeofExec)
            )
        final_string = re.sub(r"^'|'$", "", final_string)
        f = open(file, "w+")
        f.write(str(final_string))
        f.close()
        sys.stdout.flush()
        return final_string
    except Exception as e:
        raise Exception("Oops!", e)
        # TODO : Dont forget to increase coma in both if else conditions if a new column is added
        if not os.path.isfile(os.path.join(TEMP_DIR, "results", device_id + ".txt")):
            if "True" in get_network_settings:
                final_string = (
                    "status=ERROR" + ",deviceId='" + device_id + "',,,,,,,,,,"
                )
            else:
                final_string = "status=ERROR" + ",deviceId='" + device_id + "',,,,,,,"
            f = open(file, "w+")
            f.write(str(final_string))
            f.close()
        return final_string


def get_list(get_dev_list):
    """get_list"""
    # Verifies each device id based on statuses
    i = 0
    split = get_dev_list.split(";")
    command = split[0]
    status = split[1]
    in_use = split[2]
    color = split[3]
    desc = split[4]
    RESPONSE = get_device_list_response(RESOURCE_TYPE, command, status, in_use)
    DEVICE_IDS = get_device_ids(RESPONSE)
    device_list = []
    if get_handset_count(RESPONSE) > 0:
        for i in range(get_handset_count(RESPONSE)):
            device_id = DEVICE_IDS[i].firstChild.data
            device_list.append(device_id + "||" + color + "||" + desc)
            device_list = [x for x in device_list if x != 0]
        if len(device_list) > 0:
            pool_size = int(str(os.environ["pool"]))
            print("Parallel processes: " + str(pool_size))
            pool = multiprocessing.Pool(processes=pool_size, maxtasksperchild=2)
            try:
                print(
                    "\nFound " + str(len(device_list)) + " devices with status: " + desc
                )
                sys.stdout.flush()
                output = pool.map(perform_actions, device_list)
                pool.close()
                pool.terminate()
            except Exception:
                pool.close()
                pool.terminate()
                print(traceback.format_exc())
                sys.exit(-1)


def fetch_details(i, exp_number, result, exp_list):
    """ fetches details"""
    if i == exp_number:
        if "=" in result:
            exp_list = exp_list.append(result.split("=", 1)[1].replace("'", "").strip())
        else:
            exp_list = exp_list.append("-")
    return exp_list


def fig_to_base64(fig):
    img = io.BytesIO()
    plt.savefig(img, format="png", bbox_inches="tight")
    img.seek(0)
    return base64.b64encode(img.getvalue())


def print_results(results):
    """ print_results """
    i = 0
    results.sort()
    for i in range(len(results)):
        results[i] = re.sub("Results\=$", "", results[i])
        results[i] = re.sub("[,]+", "", results[i])
        if results[i]:
            if "available" in results[i] or "Reserved" in results[i]:
                print(colored(results[i], "green"))
            else:
                print(colored(results[i], "red"))
        i = i + 1


def validate_logo(logo):
    try:
        send_request(logo)
    except Exception as e:
        print("Exception: " + str(e))
        os.environ["company_logo"] = os.environ["perfecto_logo"]


def create_summary(df, title, column, name):
    fig = pl.figure(figsize=(15, 2))
    pl.suptitle(title)
    ax1 = pl.subplot(121, aspect="equal", facecolor="#fffffa")
    fig.patch.set_facecolor("white")
    fig.patch.set_alpha(1)
    try:
        df[column].value_counts().sort_index().plot(
            kind="pie",
            y="%",
            ax=ax1,
            autopct="%1.0f%%",
            startangle=30,
            shadow=False,
            labels=None,
            legend=True,
            x=df[column].unique,
            fontsize=7,
            pctdistance=1.2
        )
        pl.ylabel("")
        status = []
        for i in range(len(df[column].value_counts().sort_index().to_frame())) :
            status.append(df[column].value_counts().sort_index().to_frame().iloc[i].name)
        ax1.legend(labels=status, prop={'size': 7}, bbox_to_anchor=(1,0.5), loc="best")
        # plot table
        ax2 = pl.subplot(122, facecolor="#fffffa")
        ax2.patch.set_facecolor("white")
        ax2.patch.set_alpha(1)
        pl.axis("off")
        tbl = table(ax2, df[column].value_counts(), loc="center")
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, "results", name + ".png"))
        summary = '<img src="data:image/png;base64, {}"'.format(encoded.decode("utf-8"))
    except Exception as e:
        print("API: " + df.url + " throws " + df.msg + ". Raise a support case to get API access for fetching user list/ device list.")
        summary = '<div style="color: white;">' + df.msg + '. Raise a support case to get API access for fetching user list/ device list.</div><img '
    return summary


def prepare_graph(df, column):
    """ prepare graph """
    if column in df.columns:
        fig = pl.figure()
        fig.patch.set_facecolor("white")
        fig.patch.set_alpha(1)
        ax = (  
            df[column]
            .value_counts()
            .sort_index()
            .plot(kind="bar", fontsize=14, stacked=False, colormap='Paired', figsize=(25, 10), ylim=(0, 2))
        )
        ax.set_title(column, fontsize=20)
        ax.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.2f"))
        ax.patch.set_facecolor("white")
        ax.patch.set_alpha(0.1)
        pl.yticks(df[column].value_counts(), fontsize=10, rotation=40)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, "results", column + ".png"))
        return '<img src="data:image/png;base64, {}"'.format(encoded.decode("utf-8"))
    else:
        return "<img "


def prepare_html(user_html, table3, day, delete_reserve_df):
    """ prepare_html """
    print(colored("\nFinal Devices list:", "magenta"))
    # copies all device status to final summary
    for r, d, f in os.walk(os.path.join(TEMP_DIR, "results")):
        for file in f:
            if ".txt" in file:
                with open(os.path.join(r, file)) as f:
                    with open(os.path.join(r, "Final_Summary.txt"), "a") as f1:
                        for line in f:
                            f1.write(line)
                            f1.write("\n")
    file = os.path.join(TEMP_DIR, "results", "Final_Summary.txt")
    cloudname = os.environ["CLOUDNAME"]
    try:
        f = open(file, "r")        
        result = f.read()
        f.close()
        print_results(result.split("\n"))
        if "true" in os.environ["PREPARE_ACTIONS_HTML"]:
            results = result.split("\n")
            # TODO Add a new list and number below & a new_dict if a new column is added
            new_dict = {}
            deviceids = []
            status = []
            description = []
            manufacturer = []
            model = []
            osVersion = []
            operator = []
            phonenumber = []
            airplanemode = []
            wifi = []
            data = []
            lastExecAt = []
            action_results = []
            for result in results:
                if len(result) > 0:
                    new_result = result.split(",")
                    new_list = []
                    i = 0
                    for result in new_result:
                        fetch_details(i, 0, result, status)
                        fetch_details(i, 1, result, deviceids)
                        fetch_details(i, 2, result, manufacturer)
                        fetch_details(i, 3, result, model)
                        fetch_details(i, 4, result, osVersion)
                        fetch_details(i, 5, result, description)
                        fetch_details(i, 6, result, operator)
                        fetch_details(i, 7, result, phonenumber)
                        fetch_details(i, 8, result, airplanemode)
                        fetch_details(i, 9, result, wifi)
                        fetch_details(i, 10, result, data)
                        fetch_details(i, 11, result, action_results)
                        fetch_details(i, 12, result, lastExecAt)
                        new_list.append(result)
                        i = i + 1
            pandas.set_option("display.max_columns", None)
            pandas.set_option("display.max_colwidth", 100)
            pandas.set_option("colheader_justify", "center")
            get_network_settings = os.environ["GET_NETWORK_SETTINGS"]
            reboot = os.environ["REBOOT"]
            cleanup = os.environ["CLEANUP"]
            reserve = os.environ["RESERVE"]
            if "True" in get_network_settings or "True" in reboot or "True" in cleanup or "True" in reserve:
                new_dict = {
                    "Status": status,
                    "Device Id": deviceids,
                    "Manufacturer": manufacturer,
                    "Model": model,
                    "OS Version": osVersion,
                    "Description": description,
                    "Operator": operator,
                    "Phone number": phonenumber,
                    "AirplaneMode": airplanemode,
                    "Wifi": wifi,
                    "Data": data,
                    "Results": action_results,
                    "Last Rebooted At": lastExecAt,
                }
            else:
                new_dict = {
                    "Status": status,
                    "Device Id": deviceids,
                    "Manufacturer": manufacturer,
                    "Model": model,
                    "OS Version": osVersion,
                    "Description": description,
                    "Operator": operator,
                    "Phone number": phonenumber
                }
            df = pandas.DataFrame(new_dict)
            df = df.sort_values(by="Manufacturer")
            df = df.sort_values(by="Model")
            df = df.sort_values(by="Status")
            df.reset_index(drop=True, inplace=True)
            device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
            current_time = datetime.datetime.now().strftime("%c")
            title = (
                cloudname.upper()
                + " cloud status summary of "
                + device_list_parameters
                + os.environ["device_status"]
                + " @ "
                + current_time
            )
            summary = create_summary(df, title, "Status", "device_summary")
            plt.close("all")

            df = df.sort_values(by="Model")
            df = df.sort_values(by="Status")
        # skipping csv output as we now have full device list api response
        #         df.to_csv(os.path.join(TEMP_DIR , 'results','output.csv'), index=False)
        # Futuristic:
        #     le = preprocessing.LabelEncoder()
        #     #convert the categorical columns into numeric
        #     dfs = df.copy()
        #     encoded_value = le.fit_transform(dfs['Device Id'])
        #     dfs['Device Id'] = le.fit_transform(dfs['Device Id'])
        #     dfs['Status'] = le.fit_transform(dfs['Status'])
        #     dfs['Model'] = le.fit_transform(dfs['Model'])
        #     dfs['OS Version'] = le.fit_transform(dfs['OS Version'])
        #     dfs['Operator'] = le.fit_transform(dfs['Operator'])
        #     dfs['Phone number'] = le.fit_transform(dfs['Phone number'])
        #     if  "True" in get_network_settings or  "True" in reboot or  "True" in cleanup:
        #         dfs['AirplaneMode'] = le.fit_transform(dfs['AirplaneMode'])
        #         dfs['Wifi'] = le.fit_transform(dfs['Wifi'])
        #         dfs['Data'] = le.fit_transform(dfs['Data'])
        #         dfs['Results'] = le.fit_transform(dfs['Results'])
        #     print(dfs)
        #     cols = [col for col in dfs.columns if col not in ['Status','Phone number', 'OS Version', 'Model', 'Operator']]
        #     data = dfs[cols]
        #     target = dfs['Status']
        #     print(data)
        #     print(target)
        returnBoolean = True
    except FileNotFoundError:
        print("No devices found/ Re-check your arguments")
        summary = """<div style="color: white;">No devices matches your search criteria.</div><img '"""
        df =pandas.DataFrame([])
        returnBoolean = False

    html_string = (
        """
    <html lang="en">
        <head>
        <meta http-equiv="Content-Security-Policy" content="default-src *; img-src * data: http:; script-src 'unsafe-inline' 'unsafe-eval' *; style-src 'unsafe-inline' *">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <meta content="text/html; charset=iso-8859-2" http-equiv="Content-Type">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
        <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
                <head><title>"""
        + cloudname.upper()
        + """ Cloud Status</title>
        <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
        <script>
        $(document).ready(function(){{
            // Add smooth scrolling to all links
            $("a").on('click', function(event) {{

                // Make sure this.hash has a value before overriding default behavior
                if (this.hash !== "") {{
                // Prevent default anchor click behavior
                event.preventDefault();

                // Store hash
                var hash = this.hash;

                // Using jQuery's animate() method to add smooth page scroll
                // The optional number (800) specifies the number of milliseconds it takes to scroll to the specified area
                $('html, body').animate({{
                    scrollTop: $(hash).offset().top
                }}, 800, function(){{
            
                    // Add hash (#) to URL when done scrolling (default click behavior)
                    window.location.hash = hash;
                }});
                }} // End if
            }});
        }});
        $(document).ready(function(){{
            $("#topNav").css('display', 'inline-block');
            $("#download").css('display', 'inline-block')
            $("#download2").css('display', 'inline-block');
            $("#myInput").css('display', 'inline-block');
            $("#myInput2").css('display', 'inline-block');
            $("#myInput3").css('display', 'inline-block');
            $("#myInput4").css('display', 'inline-block');
            
            $("#myInput").on("keyup", function() {{
            var value = $(this).val().toLowerCase();
            $("#devicetable tbody tr").filter(function() {{
                $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
            }});
            }});
        }});
                    $(document).ready(function(){{
            $("#myInput2").on("keyup", function() {{
            var value = $(this).val().toLowerCase();
            $("#usertable tbody tr").filter(function() {{
                $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
            }});
            }});
        }});
        $(document).ready(function(){{
            $("#myInput3").on("keyup", function() {{
            var value = $(this).val().toLowerCase();
            $("#repotable tbody tr").filter(function() {{
                $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
            }});
            }});
        }});
        $(document).ready(function(){{
            $("#myInput4").on("keyup", function() {{
            var value = $(this).val().toLowerCase();
            $("#reservetable tbody tr").filter(function() {{
                $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
            }});
            }});
        }});
        </script>
        <script type="text/javascript">
                $(document).ready(function(){{
                $("#slideshow > div:gt(0)").show();
                $("tbody tr:contains('disconnected')").css('background-color','#fcc');
                $("tbody tr:contains('ERROR')").css('background-color','#fcc');
                $("tbody tr:contains('unavailable')").css('background-color','#fcc');
                $("tbody tr:contains('busy')").css('background-color','#fcc');
                var table = document.getElementById("devicetable");
                var rowCount = table.rows.length;
                for (var i = 0; i < rowCount; i++) {{
                    if ( i >=1){{
                    available_column_number = 0;
                    device_id_column_number = 1;
                        if (table.rows[i].cells[available_column_number].innerHTML == "available" || table.rows[i].cells[available_column_number].innerHTML == "Reserved") {{
                            for(j = 0; j < table.rows[0].cells.length; j++) {{
                                table.rows[i].cells[j].style.backgroundColor = '#e6fff0';
                                    if (table.rows[i].cells[(table.rows[0].cells.length - 1)].innerHTML.indexOf("failed") > -1) {{
                                            table.rows[i].cells[j].style.color = '#660001';
                                            table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
                                    }}
                                }}
                            var deviceId = table.rows[i].cells[device_id_column_number].innerHTML;
                            var url = 'https://"""
        + cloudname.upper()
        + """.app.perfectomobile.com/lab/devices';
                            var row = $('<tr></tr>')
                            var link = document.createElement("a");
                            link.href = url;
                            link.innerHTML = deviceId;
                            link.target = "_blank";
                            table.rows[i].cells[device_id_column_number].innerHTML = "";
                            table.rows[i].cells[device_id_column_number].appendChild(link);
                        }}else{{
                            for(j = 0; j < table.rows[0].cells.length; j++) {{
                                table.rows[i].cells[j].style.color = '#660001';
                                        table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
                            }}
                        }}
                    }}
                }}
                }});
                function myFunction() {{
                var x = document.getElementById("myTopnav");
                if (x.className === "topnav") {{
                x.className += " responsive";
                }} else {{
                x.className = "topnav";
                }}
            }}
            function zoom(element) {{
                        var data = element.getAttribute("src");
                        let w = window.open('about:blank');
                        let image = new Image();
                        image.src = data;
                        setTimeout(function(){{
                        w.document.write(image.outerHTML);
                        }}, 0);
                    }}
            function autoselect(element) {{
                    var data = element.getAttribute("id");
                    document.getElementById(data + "-1").checked = true;
            }}     
        </script>

        <meta name="viewport" content="width=device-width, initial-scale=1">
        </head>
            <style>

        html {{
            height:100%;
        }}
        
        .tabbed {{
            z-index: 0;
            display:  flex;
            text-align: left;
            flex-wrap: wrap;
            box-shadow: 0 0 20px rgba(186, 99, 228, 0.4);
            font-size: 12px;
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            }}
            .tabbed > input {{
            display: none;
            }}
            .tabbed > input:checked + label {{
            font-size: 14px;
            text-align: center;
            color: white;
            background-image: linear-gradient(to left, #bfee90, #333333, black,  #333333, #bfee90);
            }}
            .tabbed > input:checked + label + div {{
            color:darkslateblue;
            display: block;
            }}
            .tabbed > label {{
            background-image: linear-gradient(to left, #fffeea,  #333333, #333333 ,#333333 ,#333333 , #333333, #fffeea);
            color: white;
            text-align: center;
            display: block;
            order: 1;
            flex-grow: 1;
            padding: .3%;
            }}
            .tabbed > div {{
            order: 2;
            flex-basis: 100%;
            display: none;
            padding: 10px;
            overflow-x: auto;
            }}

            /* For presentation only */
            .container {{
            width: 100%;
            margin: 0 auto;
            background-color: black;
            box-shadow: 0 0 20px rgba(400, 99, 228, 0.4);
            }}

            .tabbed {{
            border: 1px solid;
            }}

            hr {{
            background-color: white;
            height: 5px;
            border: 0;
            margin: 10px 0 0;
            }}
            
            hr + * {{
            margin-top: 10px;
            }}
            
            hr + hr {{
            margin: 0 0;
            }}

        .mystyle {{
            font-size: 12pt;
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            border-collapse: collapse;
            border: 2px solid black;
            margin:auto;
            box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            background-color: #fffffa;
        }}

        .mystyle body {{
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            table-layout: auto;
            position:relative;
        }}

        #slide{{
            width:100%;
            height:auto;
        }}

        #myInput, #myInput2, #myInput3, #myInput4 {{
            background-image: url('https://cdn4.iconfinder.com/data/icons/sapphire-storm-1/32/color-web3-18-256.png');
            background-position: 2px 4px;
            background-repeat: no-repeat;
            background-size: 25px 15px;
            width: 40%;
            height:auto;
            font-weight: bold;
            font-size: 12px;
            padding: 11px 20px 12px 40px;
            box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            display:none;
            height: 20px;
        }}

        body {{
            background-color: black;
            height: 100%;
            background-repeat:  repeat-y;
            background-position: right;
            background-size:  contain;
            background-attachment: initial;
            opacity:.93;
        }}

        h4 {{
            font-family:monospace;
        }}

        @keyframes slide {{
            0% {{
            transform:translateX(-25%);
            }}
            100% {{
            transform:translateX(25%);
            }}
        }}

        .mystyle table {{
            table-layout: auto;
            width: 100%;
            height: 100%;
            position:relative;
            border-collapse: collapse;
        }}

        tr:hover {{background-color:grey;}}

        .mystyle td {{
            font-size: 12px;
            position:relative;
            padding: 5px;
            width:10%;
            color: black;
            border-left: 1px solid #333;
            border-right: 1px solid #333;
            background: #fffffa;
            text-align: center;
        }}

        table.mystyle td:first-child {{ text-align: left; }}   

        table.mystyle thead {{
            background: grey;
            font-size: 14px;
            position:relative;
            border: 1px solid black;
        }}

        table.mystyle thead th {{
            line-height: 200%;
            font-size: 14px;
            font-weight: normal;
            color: #fffffa;
            text-align: center;
            transition:transform 0.25s ease;
        }}

        table.mystyle thead th:hover {{
            -webkit-transform:scale(1.01);
            transform:scale(1.01);
        }}

        table.mystyle thead th:first-child {{
            border-left: none;
        }}

        .topnav {{
            overflow: hidden;
            background-color: black;
            opacity: 0.9;
        }}

        .topnav a {{
            float: right;
            display: block;
            color: #333333;
            text-align: center;
            padding: 12px 15px;
            text-decoration: none;
            font-size: 12px;
            position: relative;
            border: 1px solid #6c3;
            font-family: "Trebuchet MS", Helvetica, sans-serif;
        }}

        #summary{{
            box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
            position: relative;
            cursor: pointer;
            padding: .1%;
            border-style: outset;
            border-radius: 1px;
            border-width: 1px;
        }}
        
        #logo{{
            box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
            position: relative;
            cursor: pointer;
            border-style: outset;
            border-radius: 1px;
            border-width: 1px;
        }}

        .topnav a.active {{
            background-color: #333333;
            color: white;
            font-weight: lighter;
        }}

        .topnav .icon {{
            display: none;
        }}

        @media screen and (max-width: 600px) {{
            .topnav a:not(:first-child) {{display: none;}}
            .topnav a.icon {{
            color: #DBDB40;
            float: right;
            display: block;
            }}
        }}

        @media screen and (max-width: 600px) {{
            .topnav.responsive {{position: relative;}}
            .topnav.responsive .icon {{
            position: absolute;
            right: 0;
            top: 0;
            }}
            .topnav.responsive a {{
            float: none;
            display: block;
            text-align: left;
            }}
        }}

        * {{
            box-sizing: border-box;
        }}

        img {{
            vertical-align: middle;
        }}

        .containers {{
            position: relative;
        }}

        .mySlides {{
            display:flex;
            width:90%;
            padding-bottom: 3%;
        }}

        #slideshow {{
            cursor: pointer;
            margin:.01% auto;
            position: relative;
            height: 55%;
        }}

        #ps{{
            height: 10%;
            margin-top: 0%;
            margin-bottom: 90%;
            background-position: center;
            background-repeat: no-repeat;
            background-blend-mode: saturation;
        }}

        #slideshow > div {{
            position: relative;
            width: 90%;
        }}

        #download, #download2 {{
            background-color: #333333;
            border: none;
            color: white;
            font-size: 12px;
            cursor: pointer;
            display:none;
        }}

        #download:hover, #download2:hover{{
            background-color: RoyalBlue;
        }}

        .glow {{
            font-size: 15px;
            color: white;
            text-align: center;
            -webkit-animation: glow 1s ease-in-out infinite alternate;
            -moz-animation: glow 1s ease-in-out infinite alternate;
            animation: glow 1s ease-in-out infinite alternate;
        }}

        @-webkit-keyframes glow {{
            from {{
            text-shadow: 0 0 10px #fff, 0 0 20px #fff, 0 0 30px #e60073, 0 0 40px #e60073, 0 0 50px #e60073, 0 0 60px #e60073, 0 0 70px #e60073;
            }}
            
            to {{
            text-shadow: 0 0 20px #fff, 0 0 30px #ff4da6, 0 0 40px #ff4da6, 0 0 50px #ff4da6, 0 0 60px #ff4da6, 0 0 70px #ff4da6, 0 0 80px #ff4da6;
            }}
        }}

        .reportHeadingDiv {{
            background-color: #333333; 
            text-align: center;
        }}

        .reportDiv {{
            overflow-x: auto;
            align:center;
            text-align: center;
        }}

        #report{{
            box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
            overflow-x: auto;
            min-width:100%;
        }}
        
        #myTopnav{{
            display:none;
        }}
        </style>
        <body bgcolor="#FFFFED">
        <div class="topnav" id="myTopnav">
            <a href="result.html" class="active">Home</a>
            <a href="https://"""
        + cloudname.upper()
        + """.perfectomobile.com" target="_blank" class="active">"""
        + cloudname.upper()
        + """ Cloud</a>
            <a href="https://developers.perfectomobile.com" target="_blank" class="active">Docs</a>
            <a href="https://www.perfecto.io/services/professional-services-implementation" target="_blank" class="active">Professional Services</a>
            <a href="https://support.perfecto.io/" target="_blank" class="active">Perfecto Support</a>
            <a href="javascript:void(0);" aria-label="first link" class="icon" onclick="myFunction()">
            <i class="fa fa-bars"></i>
            </a>
        </div>

        <div style="text-align: center">
            
            <div class="container">
                <div class="tabbed">
                    <input type="radio" id="tabbed-tab-1-1" onClick='autoselect(this)' name="tabbed-tab-1"><label for="tabbed-tab-1-1">Users</label>
                    <div>
                        <div class="tabbed">
                            <input type="radio" id="tabbed-tab-1-1-1" name="tabbed-tab-1-1"><label for="tabbed-tab-1-1-1">List</label>
                            <div align="center">
                            
                            <a href="https://"""
        + cloudname.upper()
        + """.perfectomobile.com" target="_blank" class="site-logo">
                            <img id="logo" src="""
        + os.environ["company_logo"]
        + """ style="margin:1%;" alt="Company logo" ></a> 
        <div class="reportDiv"></p><br>
                            """
        + create_summary(user_html, "Users list Status", "status", "user_summary")
        + """ alt='user_summary' id='summary' onClick='zoom(this)'></img></br></p></div>
                            <input id="myInput2" aria-label="search" type="text" placeholder="Search..">&nbsp;&nbsp;&nbsp;
                            <a id ="download" href="./get_users_list.xlsx" aria-label="A link to users .xlsx file is present." class="btn"><i class="fa fa-download"></i> Users List</a>
                            </br> </br>
                            <div style="overflow-x:auto;">
                                {table2}
                            </div>
                            
                            </div>
                            </div>
                    </div>
                <input type="radio" id="tabbed-tab-1-2" onClick='autoselect(this)' name="tabbed-tab-1" checked><label id="tabbed-device" for="tabbed-tab-1-2">Device</label>
                <div>
                    <div class="tabbed">
                        <input type="radio" id="tabbed-tab-1-2-1" name="tabbed-tab-1-1" checked><label for="tabbed-tab-1-2-1">List</label>
                        <div  align="center">
                        
                            <a href="https://"""
        + cloudname.upper()
        + """.perfectomobile.com" target="_blank" class="site-logo">
                            <img id="logo" src="""
        + os.environ["company_logo"]
        + """ style="margin:1%;" alt="Company logo" ></a> 
        <div style="overflow-x:auto;">
                                {table4}</br></p>
                            </div>
                            <div class="reportDiv">"""
        + summary
        + """ alt='summary' id='summary' onClick='zoom(this)'></img> </br></p></div>
                            <input id="myInput" aria-label="search" type="text" placeholder="Search..">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                <a id ="download2" href="./get_devices_list.xlsx" aria-label="A link to a .xlsx file is present." class="btn"><i class="fa fa-download"></i> Full Devices List</a>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                </br> </br>
                                    <div style="overflow-x:auto;">
                                        {table}
                                    </div>
                                </br>
                    </div>
                    <input type="radio" id="tabbed-tab-1-2-2" name="tabbed-tab-1-1"><label for="tabbed-tab-1-2-2">Graphs</label>
                        <div align="center">

                                <div style="overflow-x:auto;height:90%">
                                <div class="containers" align="center" id = "slideshow">
                                    <div class="w3-content w3-section"  style="max-width:90%; max-height:90%;height:90%;width:90%;">
                                    """
        + prepare_graph(df, "Manufacturer")
        + """ alt="Device Status" class="mySlides"  onClick='zoom(this)' id="slide">
                                    """
        + prepare_graph(df, "Model")
        + """ alt="Model" class="mySlides"  onClick='zoom(this)' id="slide">
                                    """
        + prepare_graph(df, "OS Version")
        + """ alt="Version" class="mySlides" onClick='zoom(this)' id="slide">
                                    """
        + prepare_graph(df, "Operator")
        + """ alt="Operator" class="mySlides" onClick='zoom(this)' id="slide">
                                    """
        + prepare_graph(df, "Description")
        + """ alt="Description" class="mySlides"  onClick='zoom(this)' id="slide">
                                    </div>
                                </div>
                            </div>
                                
                        </div>
                </div>
            </div>
                    {table3}
                </div>
                <a target="_blank" style="font-size:12;font-family:"Trebuchet MS", Helvetica, sans-serif;color:powderblue;" href="https://clearbit.com">Logos provided by Clearbit</a>
            </div>
            <script>
            var myIndex = 0;
            carousel();

            function carousel() {{
            var i;
            var x = document.getElementsByClassName("mySlides");
            for (i = 0; i < x.length; i++) {{
                x[i].style.display = "none";
            }}
            myIndex++;
            if (myIndex > x.length) {{myIndex = 1}}
            x[myIndex-1]topnavdisplay = "block";
            setTimeout(carousel, 2000); // Change image every 2 seconds
            }}
            </>
        </body>
    </html>
    """
    )

    # OUTPUT AN HTML FILE
    clean_repo = os.environ["clean_repo"]
    with open(os.path.join(TEMP_DIR, "output", "result.html"), "w") as f:
        if "NA" != clean_repo:
            heading = (
                """<input type="radio" onClick='autoselect(this)' id="tabbed-tab-1-3" name="tabbed-tab-1"><label for="tabbed-tab-1-3">Repository</label>
                        <div>
                        <div class="tabbed">
                            <input type="radio" id="tabbed-tab-1-3-1" name="tabbed-tab-1-1" ><label for="tabbed-tab-1-3-1">List</label>
                            <div align="center">
                        <b><h4><p style="color:white">"""
                + cloudname.upper()
                + """ Media Repository Cleanup Status for files older than """
                + str(day)
                + """ days: Total - """
                + str(table3.shape[0])
                + """</p></font></h4>
                            <input id="myInput3" aria-label="search" type="text" placeholder="Search.."></br> </br>
                        <div style="overflow-x:auto;"></b>"""
            )
            f.write(
                html_string.format(
                    table=df.to_html(
                        classes="mystyle", table_id="devicetable", index=False
                    ),
                    table2=user_html.to_html(
                        classes="mystyle",
                        table_id="usertable",
                        justify="justify-all",
                        index=False,
                    ),
                    table3=heading
                    + table3.to_html(
                        classes="mystyle", table_id="repotable", index=False
                    )
                    + "</div></div></div></div></br>",
                    table4=delete_reserve_df,
                )
            )
        else:
            try:
                f.write(
                    html_string.format(
                        table=df.to_html(
                            classes="mystyle", table_id="devicetable", index=False
                        ),
                        table2=user_html.to_html(
                            classes="mystyle",
                            table_id="usertable",
                            justify="justify-all",
                            index=False,
                        ),
                        table3="",
                        table4=delete_reserve_df,
                    )
                )
            except:
                f.write(
                    html_string.format(
                        table=df.to_html(
                            classes="mystyle", table_id="devicetable", index=False
                        ),
                        table2="",
                        table3="",
                        table4=""
                    )
                )
    webbrowser.open(
        "file://" + os.path.join(TEMP_DIR, "output", "result.html"), new=0
    )
    print("Results: file://" + os.path.join(TEMP_DIR, "output", "result.html"))
    return returnBoolean


def send_request_repo(url, content):
    try:
        response = urllib.request.urlopen(url.replace(" ", "%20"))
    except Exception as e:
        return e
    sys.stdout.flush()
    return response


def send_request_for_repository(url, content, key):
    map = []
    response = send_request_repo(url, content)

    # if "500" in str(response):
    #     raise RuntimeError(
    #         "Failed to list repository items - Repository item: "
    #         + key
    #         + "  was not found in media repository, url:"
    #         + str(url)
    #     )
    #     sys.exit(-1)
    text = response.read().decode("utf-8")
    map = json.loads(text)
    return map


def getActualDate(map):
    # date is fetched here
    try:
        date = map["item"]["creationTime"]["formatted"]
    except KeyError:
        return ""
    dateOnly = date.split("T")
    return datetime.datetime.strptime(dateOnly[0], "%Y-%m-%d")


def getPastDate(days):
    # Logic for fetching past days based on user preference
    today = datetime.datetime.today()
    pastDate = timedelta(days=int(float(days)))
    return today - pastDate

def sendAPI(resource_type, resource_key, operation):
    url = get_url(str(resource_type), resource_key, operation)
    #new repo doesnt seem to require admin for now
    map = send_request_for_repository(url, "", resource_key)
    if "completionCode" in map:
        if operation == "delete" and map["completionCode"] == "notOwner":
            admin = os.environ["repo_admin"]
            if "true" in admin.lower():
                url += "&admin=" + "true"
            map = send_request_for_repository(url, "", resource_key)
    return map

def deleteReserveAPI(resource_type, resource_key, operation, admin):
    url = get_url(str(resource_type), resource_key, operation)
    if "true" in admin.lower():
        url += "&admin=" + "true"
    print("Delete API url: \n" + url)
    map = send_request_for_repository(url, "", operation)
    return map

def sendAPI_repo(resource_type, resource_key, operation):
    url = get_url(str(resource_type), "", operation)
    print("\nRepository API Raw url: \n" + url)
    return send_request_for_repository(url, "", resource_key)


def fetch_details_repo(i, exp_number, result, exp_list):
    """ fetches details"""
    if i == exp_number:
        if ":" in result:
            exp_list = exp_list.append(result.split(":", 1)[1].replace("'", "").strip())
        else:
            exp_list = exp_list.append("-")
    return exp_list


def run_commands(value):
    # Get date of repository items
    FINAL_LIST = []
    DAYS = os.environ["repo_days"]
    DELETE = os.environ["repo_delete"]
    map = sendAPI(os.environ["repo_resource_type"], value, "info")
    actualDate = getActualDate(map)
    if not (str(actualDate) == ""):
        expectedDate = getPastDate(DAYS)
        expDate = str.split(str(expectedDate), " ")
        actDate = str(str.split(str(actualDate), " ")[0])
        # DELETES the item if older than expected date
        if actualDate < expectedDate:
            print(
                colored(
                    "File: "
                    + value
                    + " with actual creation date: "
                    + actDate
                    + " was created before "
                    + str(DAYS)
                    + " days.",
                    "red",
                )
            )
            # DELETE item from the repository
            if DELETE.lower() == "true":
                map = sendAPI(os.environ["repo_resource_type"], value, "delete")
                print("Delete API output of item: " + str(value) + ":\n " + str(map))
                status = map["status"]
                if status != "Success":
                    FINAL_LIST.append(
                        "File:"
                        + value
                        + ";Created on:"
                        + actDate
                        + ";Comparison:is older than;Days:"
                        + DAYS
                        + ";Deleted?:Unable to delete!;"
                    )
                    raise RuntimeError("Repository item " + value + " was not deleted. API Response:" + str(map))
                    sys.exit(-1)
                else:
                    FINAL_LIST.append(
                        "File:"
                        + value
                        + ";Created on:"
                        + actDate
                        + ";Comparison:is older than;Days:"
                        + DAYS
                        + ";Deleted?:Yes;"
                    )
            else:
                FINAL_LIST.append(
                    "File:"
                    + value
                    + ";Created on:"
                    + actDate
                    + ";Comparison:is older than;Days:"
                    + DAYS
                    + ";Deleted?:No;"
                )
        else:
            print(
                colored(
                    "File: "
                    + value
                    + " with actual creation date: "
                    + actDate
                    + " was created within the last "
                    + str(DAYS)
                    + " days.",
                    "green",
                )
            )
    #                   FINAL_LIST.append('File:' + value + ';Created on:' + actDate + ';Comparison:is younger than;Days:' + DAYS + ';Deleted?:No;')
    fileName = uuid.uuid4().hex + ".txt"
    file = os.path.join(TEMP_DIR, "repo_results", fileName)
    f = open(file, "w+")
    f.write(str(FINAL_LIST))
    f.close()


def manage_repo(resource_key):
    # Get list of repository items
    map = sendAPI_repo(os.environ["repo_resource_type"], resource_key, "list")
    try:
        itemList = map["items"]
        itemList = [x for x in itemList if x.startswith(resource_key)] 
        sys.stdout.flush()
        print("Respository items list: " + str(itemList))
    except Exception as e:
        print(str(e))
        sys.exit(-1)
    # debug
    #     for value in itemList:
    #         run_commands(value)
    pool_size = multiprocessing.cpu_count() * 2
    repo_folder_pool = multiprocessing.Pool(processes=pool_size, maxtasksperchild=2)
    try:
        FINAL_LIST = repo_folder_pool.map(run_commands, itemList)
        repo_folder_pool.close()
        repo_folder_pool.terminate()
    except Exception:
        repo_folder_pool.close()
        repo_folder_pool.terminate()
        print(traceback.format_exc())
        sys.exit(-1)


def deleteOlderFiles(resource_type, delete, admin, repo_paths, days):
    os.environ["repo_delete"] = delete
    os.environ["repo_days"] = days
    os.environ["repo_resource_type"] = resource_type
    os.environ["repo_admin"] = admin
    create_dir(os.path.join(TEMP_DIR, "repo_results"), True)
    I = 0
    REPO_LIST = repo_paths.split(",")
    # debug:
    #     for repo in REPO_LIST:
    #         manage_repo(repo)
    procs = []
    for li in REPO_LIST:
        proc = Process(target=manage_repo, args=(str(li),))
        procs.append(proc)
        proc.start()
    try:
        for proc in procs:
            proc.join()
        for proc in procs:
            proc.terminate()
    except Exception:
        proc.terminate()
        print(traceback.format_exc())
        sys.exit(-1)

    for r, d, f in os.walk(os.path.join(TEMP_DIR, "repo_results")):
        for file in f:
            if ".txt" in file:
                with open(os.path.join(r, file)) as f:
                    with open(os.path.join(r, "Final_Repo.txt"), "a") as f1:
                        for line in f:
                            f1.write(line)
                            f1.write("\n")
    file = os.path.join(TEMP_DIR, "repo_results", "Final_Repo.txt")
    try:
        f = open(file, "r")
        result = f.read()
        f.close()
    except FileNotFoundError:
        result = ""
        pass
    
    FINAL_LIST = result.split("\n")

    file = []
    created = []
    comparison = []
    days = []
    deleted = []
    final_dict = {}
    for lists in FINAL_LIST:
        if lists is not None:
            if len(lists) > 0:
                new_result = str(lists).split(";")
                i = 0
                for result in new_result:
                    if "Deleted?:" in result:
                        fetch_details_repo(
                            i,
                            new_result.index(result, i),
                            str(result).replace("]", ""),
                            deleted,
                        )
                    if "File:" in result:
                        fetch_details_repo(i, new_result.index(result, i), result, file)
                    if "Created on:" in result:
                        fetch_details_repo(
                            i, new_result.index(result, i), result, created
                        )
                    if "Comparison:" in result:
                        fetch_details_repo(
                            i, new_result.index(result, i), result, comparison
                        )
                    if "Days:" in result:
                        fetch_details_repo(i, new_result.index(result, i), result, days)
                    i = i + 1
    pandas.set_option("display.max_columns", None)
    pandas.set_option("display.max_colwidth", 100)
    pandas.set_option("colheader_justify", "center")
    final_dict = {
        "File": file,
        "Created On": created,
        "Comparison": comparison,
        "Days": days,
        "Deleted?": deleted,
    }
    df = pandas.DataFrame(final_dict)
    df = df.sort_values(by="File")
    df.style.set_properties(**{"text-align": "left"})
    sys.stdout.flush()
    return df

def parse_XML(xml, df_cols): 
    xtree = ETree.parse(xml)
    xroot = xtree.getroot()
    rows = []
    for node in xroot: 
        res = []
        res.append(node.attrib.get(df_cols[0]))
        for el in df_cols[1:]: 
            if node is not None and node.find(el) is not None:
                res.append(node.find(el).text)
            else: 
                res.append(None)
        rows.append({df_cols[i]: res[i] 
                     for i, _ in enumerate(df_cols)})  
    return pandas.DataFrame(rows, columns=df_cols)

def reserve_limiter(resource_type, limit, list, end, admin, skip_list, include_list, delete):
    user = 'reservedTo'
    deviceId = 'resourceId'
    reservationid = 'id'
    startTime = 'startTime/millis'
    endTime = 'endTime/millis'
    status = 'status'
    cols = [user, deviceId, reservationid, status, startTime, endTime]  
    response = get_reservations_list_response(resource_type, list, end, admin)
    print("Reservations list API response: \n" + response)
    executions = json.loads(response)
    count = len(executions['reservations'])
    label = ""
    if(str(delete).lower() == "true"):
        label = "List of deleted reservations"
    else:
        label = "List of reservation violations"
    html = """<div class="tabbed">
          <label for="tabbed-tab-1-2-3" style="font-size: 17px;background-image:linear-gradient(to left, #bfee90, #333333, black,  #333333, #bfee90) !important;">""" + label + """ beyond """ + limit + """ per user</label></div>
          <div style="color: white;font-size: 15px;">"""
    if count > 0 :
        ori_df = pandas.DataFrame([flatten_json(x) for x in executions['reservations']],index=range(count),columns=cols)
        ori_df[startTime] = pandas.to_datetime(ori_df[startTime], unit="ms")
        ori_df[startTime] = (
            ori_df[startTime].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        ori_df[startTime] = ori_df[startTime].dt.strftime("%d/%m/%Y %H:%M:%S")
        ori_df[endTime] = pandas.to_datetime(ori_df[endTime], unit="ms")
        ori_df[endTime] = (
            ori_df[endTime].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        ori_df[endTime] = ori_df[endTime].dt.strftime("%d/%m/%Y %H:%M:%S")
        print(ori_df)
        pandas.set_option("display.max_columns", None)
        pandas.set_option("display.max_colwidth", 100)
        pandas.set_option("colheader_justify", "center")
        deleted_df = pandas.DataFrame(columns=cols)
        i = 0
        for user_name in ori_df[user].unique():
            if(user_name in include_list.split(",") or include_list == ""):
                pandas.options.mode.chained_assignment = None
                df = ori_df[ori_df[user] == user_name] 
                df = df[df[status].isin(['SCHEDULED', 'STARTED'])] 
                df=df.sort_values([startTime,endTime])
                df[startTime] = pandas.to_datetime(df[startTime])
                df[endTime] = pandas.to_datetime(df[endTime])
                # reducing a second from endtime so that it doesnt create an unexpected overlap situation.
                df[endTime] = df[endTime] - timedelta(seconds=1)
                temp_df = df[[startTime,endTime]]
                print("\n full table of user: " + str(user_name))
                print(temp_df)
                
                temp_df = temp_df.melt(var_name = 'status',value_name = 'time').sort_values('time')
                # if there are more than expected violations, delete the reservation.
                while(temp_df['time'].shape[0] > 0):
                    if(temp_df['time'].shape[0] > 1):
                        temp_df['counter'] = np.where(temp_df['status'].eq(startTime),1,-1).cumsum()
                        # temp_df['counter'] = temp_df['status'].map({startTime:1,endTime:-1}).cumsum()
                        temp_df = temp_df[temp_df['counter'] > int(str(limit))]  
                    temp_df = temp_df[temp_df['status'] == startTime].sort_values(['time'], ascending=[False]).sort_values(['counter'], ascending=[False])
                    print((temp_df))
                    if(temp_df['time'].shape[0] > 0):
                        new_startTime = temp_df['time'].iloc[0]
                        new_df = df[df[startTime] == new_startTime]
                        new_df = new_df.sort_values([endTime], ascending=[False])
                        print("\n violation for user: " + str(user_name) +  "\n")
                        print(new_df)
                        if(temp_df['counter'].iloc[0] > (int(str(limit)) )):       
                            bool = False              
                            while(new_df.shape[0] > 0) :
                                time_range = DateTimeRange( new_df[startTime].iloc[0],  new_df[endTime].iloc[0])
                                count = 0
                                if(new_df.shape[0] > 0):
                                    df = df.reset_index(drop=True)
                                    for index, row in df.iterrows():
                                        x = DateTimeRange(df[startTime].iloc[index], df[endTime].iloc[index])
                                        if(time_range.is_intersection(x)):
                                            count+=1
                                else:
                                    break;
                                if(count > (int(str(limit)))):
                                    reservation_id = str(new_df[reservationid].iloc[0])
                                    if(new_df[user].iloc[0] not in skip_list.split(",")):
                                        if(str(delete).lower() == "true"):
                                            # delete each reservation
                                            print("id: " + reservation_id + " will be deleted now.")
                                            map = deleteReserveAPI(RESOURCE_TYPE_RESERVATIONS, reservation_id, "delete", "true")
                                            print("Delete API output of id: " + reservation_id + "\n " + str(map))
                                            try:
                                                delete_status = map["status"]
                                                if("Success" in delete_status):  
                                                    pass
                                                else:
                                                    raise Exception("Unable to delete reservation: " + reservation_id + " Response: " + map.items())    
                                            except Exception as e:
                                                raise Exception("Unable to delete reservation: " + reservation_id + " Response: " + map['errorMessage'])  
                                        deleted_df.loc[i] = new_df.iloc[0] 
                                        i=i+1
                                        ori_df.drop(ori_df[ori_df[reservationid] == reservation_id].head(1).index, inplace = True) 
                                    df.drop(df[df[reservationid] == reservation_id].head(1).index, inplace = True) 
                                    new_df.drop(new_df[new_df[reservationid] == reservation_id].head(1).index, inplace=True)
                                else:
                                    break
                            temp_df.drop(temp_df[temp_df['time'] == new_startTime].head(1).index, inplace = True) 
        
        print("\n remaining  reservations:\n")
        ori_df[endTime] = pandas.to_datetime(ori_df[endTime], format="%d/%m/%Y %H:%M:%S")
        ori_df[endTime] = ori_df[endTime] + timedelta(seconds=1) # increasing a second to match as per reservation UI
        print(ori_df.to_string(index=False))
        if(deleted_df.shape[0] > 0):
            deleted_df = deleted_df.drop(status, axis=1)
            deleted_df[endTime] = deleted_df[endTime] + timedelta(seconds=1)
            if(str(delete).lower() == "true"):
                print("\n Deleted reservations:")
            else:
                print("\n The following reservations will be deleted:")
            print(deleted_df.to_string(index=False))
            deleted_df = deleted_df.sort_values(by=user)
            deleted_df.style.set_properties(**{"text-align": "left"})
            deleted_df.columns = ['Reserved To', 'Device ID', 'Reservation Id',	'Start time', 'End time']
            deleted_df = deleted_df.drop('Reservation Id', 1)
            sys.stdout.flush()
            html += """</p><input id="myInput4" aria-label="search" type="text" placeholder="Search.."></p> """
            html += deleted_df.to_html(
                            classes="mystyle", table_id="reservetable", index=False)
        else:
            
            html += """There were no violations in reservations!"""
    else:
        html += """No matching reservations were found!"""
    html += "</div>"
    return html

def create_dir(directory, delete):
    """
    create Dir
    """
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
        else:
            if delete:
                shutil.rmtree(directory)
                os.makedirs(directory)
    except Exception as e:
        print(colored(e, "red"))
        sys.exit(-1)


def main():
    """
    Runs the perfecto actions and reports
    """
    try:
        start_time = time.time()
        delete_reserve_df = ""
        freeze_support()
        # init()
        #     """fix Python SSL CERTIFICATE_VERIFY_FAILED"""
        if not os.environ.get("PYTHONHTTPSVERIFY", "") and getattr(
            ssl, "_create_unverified_context", None
        ):
            ssl._create_default_https_context = ssl._create_unverified_context
        parser = argparse.ArgumentParser(description="Perfecto Actions Reporter")
        parser.add_argument(
            "-c",
            "--cloud_name",
            metavar="cloud_name",
            help="Perfecto cloud name. (E.g. demo)",
        )
        parser.add_argument(
            "-s",
            "--security_token",
            metavar="security_token",
            type=str,
            help="Perfecto Security Token/ Pass your Perfecto's username and password in user:password format",
        )
        parser.add_argument(
            "-d",
            "--device_list_parameters",
            metavar="device_list_parameters",
            type=str,
            help="Perfecto get device list API parameters to limit device list. Support all API capabilities which selects devices based on reg ex/strings,  Leave it empty to select all devices",
            nargs="?",
        )
        parser.add_argument(
            "-u",
            "--user_list_parameters",
            metavar="user_list_parameters",
            type=str,
            help="Perfecto get user list API parameters to limit user list. Support all API capabilities which selects users based on applicable parameters,  Leave it empty to select all users",
            nargs="?",
        )
        parser.add_argument(
            "-t",
            "--device_status",
            type=str,
            metavar="Different types of Device Connection status",
            help="Different types of Device Connection status. Values: all. This will showcase all the device status like available, disconnected, unavailable, notavailable & busy. Note: Only Available devices will be shown by default",
            nargs="?",
        )
        parser.add_argument(
            "-a",
            "--actions",
            metavar="actions",
            type=str,
            help="Perfecto actions seperated by semi-colon. E.g. reboot:true;cleanup:true;get_network_settings:true",
            nargs="?",
        )
        parser.add_argument(
            "-r",
            "--refresh",
            type=str,
            metavar="refresh",
            help="Refreshes the page with latest device status as per provided interval in seconds",
            nargs="?",
        )
        parser.add_argument(
            "-o",
            "--output",
            type=str,
            metavar="output in html",
            help="output in html. Values: true/false. Default is true",
            nargs="?",
        )
        parser.add_argument(
            "-l",
            "--logo",
            type=str,
            metavar="shows customer logo",
            help="shows client logo if valid official client website url is specified in this sample format: www.perfecto.io",
            nargs="?",
        )
        parser.add_argument(
            "-p",
            "--pool",
            type=str,
            metavar="pool",
            help="Overrides default parallel processes count",
            nargs="?",
        )
        args = vars(parser.parse_args())
        if not args["cloud_name"]:
            parser.print_help()
            parser.error(
                "cloud_name parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoactions -c demo"
            )
            exit
        if not args["security_token"]:
            parser.print_help()
            parser.error(
                "security_token parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoactions -c demo -s <<TOKEN>> || perfectoactions -c demo -s <<user>>:<<password>>"
            )
            exit
        print(str(args["cloud_name"]))
        test_list = ["perfectomobile.com","https://",".app."] 
        if [ele for ele in test_list if(ele in str(args["cloud_name"]).lower())]:
            raise Exception(
                "Kindly provide only cloud name. E.g. cloud name of demo.perfectomobile.com is just 'demo'. Hence the correct way to pass cloud name is -c demo "
            )
        os.environ["CLOUDNAME"] = args["cloud_name"]
        os.environ["TOKEN"] = args["security_token"]
        if args["device_list_parameters"]:
            device_list_parameters = args["device_list_parameters"]
        else:
            device_list_parameters = "All devices"
        os.environ["DEVICE_LIST_PARAMETERS"] = device_list_parameters
        os.environ["USER_LIST_PARAMETERS"] = "All users"
        if args["user_list_parameters"]:
            os.environ["USER_LIST_PARAMETERS"] = args["user_list_parameters"]
        os.environ[
            "perfecto_logo"
        ] = "https://logo.clearbit.com/www.perforce.com?size=120"
        if args["logo"]:
            if str("www.").lower() not in str(args["logo"]).lower():
                raise Exception(
                    "Kindly provide valid client website url. Sample format: www.perfecto.io"
                )
                sys.exit(-1)
            new_logo = "https://logo.clearbit.com/" + args["logo"] + "?size=120"
            validate_logo(new_logo)
            os.environ["company_logo"] = new_logo
        else:
            os.environ["company_logo"] = os.environ["perfecto_logo"]
        os.environ["pool"] =  str(multiprocessing.cpu_count() * 1)
        if args["pool"]:
            os.environ["pool"] = args["pool"] 
        os.environ["GET_NETWORK_SETTINGS"] = "False"
        reboot = "False"
        cleanup = "False"
        os.environ["RESERVE"] = "False"
        start_execution = "False"
        clean_repo = "NA"
        if args["actions"]:
            if "get_network_settings:true" in args["actions"]:
                os.environ["GET_NETWORK_SETTINGS"] = "True"
            if "reboot:true" in args["actions"]:
                reboot = "True"
            if "cleanup:true" in args["actions"]:
                cleanup = "True"
            if "reserve:" in args["actions"]:
                reserve = args["actions"]
                try:
                    reserve = reserve.split("reserve:")[1]
                    os.environ["RESERVE"] = "True"
                    try:
                        if (int(reserve.split(":")[0].split(";")[0])):
                            os.environ["RESERVE_TIME"] = reserve.split(":")[0].split(";")[0]
                            print("Reserve time:" + os.environ["RESERVE_TIME"])
                    except:
                        print("Wrong reservation time. defaulting to 15.")
                        os.environ["RESERVE_TIME"] = "15"
                        pass
                except IndexError as e:
                    raise Exception( "Begin with reserve parameter with a colon seperator followed by reservation time in minutes. Minimum default time is 15 minutes. E.g. reserve:15 ")
            if "reserve_limit" in args["actions"]:
                print("args:"  + str(args))
                reserve_limit = args["actions"]
                # reserve_limit = "reserve_limit:1|false||genesist@perfectomobile.com,hadass@perfectomobile.com|1"
                os.environ["reserve_limit_days"] = ""
                try:
                    reserve_limit = reserve_limit.split("reserve_limit:")[1]
                    if (int(reserve_limit.split(":")[0].split("|")[0].split(";")[0])):
                        os.environ["reserve_limit_user"] = reserve_limit.split(":")[0].split("|")[0].split(";")[0]
                        print("reserve limit user: " + os.environ["reserve_limit_user"])
                except:
                    raise Exception( "Begin with reserve_limit parameter with a colon seperator followed by reservation limit. E.g: reserve_limit:2 will limit reservations to 2 per user at any point in time." )
                try:
                    if(len(reserve_limit.split(":")[0].split("|")) > 1):
                        if(str(reserve_limit.split(":")[0].split("|")[1].split(";")[0]).lower() =="true"):
                            delete = reserve_limit.split(":")[0].split("|")[1].split(";")[0]
                        else:
                            delete = "false"
                    else:
                        delete = "false"
                except:
                    raise Exception( "Provide true to delete the reservations. default is false. E.g: reserve_limit:2|true will delete reservations to 2 per user at any point in time." )
                try:
                    if(len(reserve_limit.split(":")[0].split("|")) > 2):
                        os.environ["reserve_limit_skip"] = reserve_limit.split(":")[0].split("|")[2].split(";")[0]
                        print("reserve limit skip users: " + os.environ["reserve_limit_skip"])
                    else:
                        os.environ["reserve_limit_skip"] = ""
                except:
                    raise Exception( "Provide optional users list which needs to be skipped while limiting reservation with , seperator. E.g: reserve_limit:2|true|email@abc.com,cmain@abc.com will limit reservations to 2 and exclude email@abc.com,cmain@abc.com only")
                try:
                    if(len(reserve_limit.split(":")[0].split("|")) > 3):
                            include_list = reserve_limit.split(":")[0].split("|")[3].split(";")[0]
                            print("reserve limit include users: " + include_list)
                    else:
                        include_list = ""
                except:
                    raise Exception("Provide list of users that should be included while limiting reservation with , seperator. E.g: reserve_limit:2|true||email@abc.com,cmain@abc.com will limit reservations to 2 and include email@abc.com,cmain@abc.com only" )
                try:
                    if(len(reserve_limit.split(":")[0].split("|")) > 4):
                        os.environ["reserve_limit_days"] = reserve_limit.split(":")[0].split("|")[4].split(";")[0]
                        print("reserve limit days: " + os.environ["reserve_limit_days"])
                    else:
                        os.environ["reserve_limit_days"] = ""
                except:
                    raise Exception( "Provide optional reserve limit in days. E.g: reserve_limit:2|true|||4 will limit reservations to 2 per user beginning now to within the next 4 days" )
                
                delete_reserve_df = reserve_limiter(RESOURCE_TYPE_RESERVATIONS, os.environ["reserve_limit_user"], "list", os.environ["reserve_limit_days"], "true", os.environ["reserve_limit_skip"], include_list, delete)
            if "clean_repo" in args["actions"]:
                clean_repo = args["actions"]
            else:
                os.environ["clean_repo"] = "NA"
        os.environ["clean_repo"] = clean_repo
        # manage repo:
        clean_repo = os.environ["clean_repo"]
        if "NA" != clean_repo:
            try:
                clean_repo_var = clean_repo.split("|")
                if ";" in str(clean_repo_var[4]):
                    day = str(clean_repo_var[4]).split(";")[0]
                else:
                    day = str(clean_repo_var[4])
                repo_html = deleteOlderFiles(
                    REPOSITORY_RESOURCE_TYPE,
                    clean_repo_var[1],
                    clean_repo_var[2],
                    clean_repo_var[3],
                    day,
                )
            except Exception as e:
                raise Exception(
                    "Error: "
                    + str(e)
                )
                sys.exit(-1)
        os.environ["CLEANUP"] = cleanup
        os.environ["REBOOT"] = reboot
        if (
            "True" in os.environ["GET_NETWORK_SETTINGS"]
            or "True" in reboot
            or "True" in cleanup
        ):
            start_execution = "True"
        os.environ["START_EXECUTION"] = start_execution
        os.environ["PREPARE_ACTIONS_HTML"] = "true"
        if args["output"]:
            if "false" in str(args["output"]).lower():
                os.environ["PREPARE_ACTIONS_HTML"] = "false"
        os.environ["perfecto_actions_refresh"] = "false"
        if args["refresh"]:
            if int(args["refresh"]) >= 0:
                os.environ["perfecto_actions_refresh"] = args["refresh"]
        # create results path and files
        create_dir(os.path.join(TEMP_DIR, "results"), True)
        create_dir(os.path.join(TEMP_DIR, "repo_results"), True)
        create_dir(os.path.join(TEMP_DIR, "output"), True)
        # result = get_xml_to_xlsx(RESOURCE_TYPE, "list", 'get_devices_list.xlsx')
        # get device list to excel

        devlist = Pool(processes=1)
        try:
            result = devlist.apply_async(
                get_xml_to_xlsx, [RESOURCE_TYPE, "list", "get_devices_list.xlsx"]
            )
        except Exception:
            devlist.close()
            print(traceback.format_exc())
            sys.exit(-1)
        # cradleList = Pool(processes=1)
        # try:
        #     result = cradleList.apply_async(
        #         get_xml_to_xlsx, [RESOURCE_TYPE_CRADLES, "list", "get_cradles_list.xlsx"]
        #     )
        # except Exception:
        #     cradleList.close()
        #     print(traceback.format_exc())
        #     sys.exit(-1)
        # cradleInfo = Pool(processes=1)
        # try:
        #     result = cradleInfo.apply_async(
        #         get_cradles_xml_to_xlsx, [RESOURCE_TYPE_CRADLES, "info", "get_cradles_info.xlsx"]
        #     )
        # except Exception:
        #     cradleInfo.close()
        #     print(traceback.format_exc())
        #     sys.exit(-1)
        #Dont add the below to threading as it will freeze if issues found.
        user_html = get_json_to_xlsx(RESOURCE_TYPE_USERS, "list", 'get_users_list.xlsx')
        get_dev_list = []
        if args["device_status"]:
            os.environ["device_status"] = " with " + args["device_status"].lower() + " status"
            if "all" in args["device_status"].lower():
                #             may require for debug single threads
                #             get_list("list;connected;false;green;available")
                #             get_list("list;connected;true;red;busy")
                #             get_list("list;disconnected;;red;disconnected")
                #             get_list("list;unavailable;;red;unavailable")
                get_dev_list = [
                    "list;connected;true;red;busy",
                    "list;disconnected;;red;disconnected",
                    "list;unavailable;;red;unavailable",
                    "list;connected;false;green;available",
                ]
            elif "disconnected".lower() in args["device_status"].lower():
                get_dev_list = [
                    "list;disconnected;;red;disconnected"
                ]
            elif "unavailable".lower() in args["device_status"].lower():
                get_dev_list = [
                   "list;unavailable;;red;unavailable"
                ]
            elif "notavailable" in args["device_status"].lower():
                get_dev_list = [
                    "list;disconnected;;red;disconnected",
                    "list;unavailable;;red;unavailable"
                ]
                
            try:
                procs = []
                for li in get_dev_list:
                    proc = Process(target=get_list, args=(str(li),))
                    procs.append(proc)
                    proc.start()
                for proc in procs:
                    proc.join()
                for proc in procs:
                    proc.terminate()
            except Exception:
                    print(traceback.format_exc())
                    proc.terminate()    
                    sys.exit(-1)
        else:
            os.environ["device_status"] = ""
            if not args["device_list_parameters"]:
                os.environ["DEVICE_LIST_PARAMETERS"] = "Available Devices only"
            get_list("list;connected;false;green;available")
        if "NA" != clean_repo:
            bool = prepare_html(user_html, repo_html, day, delete_reserve_df)
        else:
            bool = prepare_html(user_html, "", "", delete_reserve_df)
        print("--- Completed in : %s seconds ---" % (time.time() - start_time))
        # Keeps refreshing page with expected arguments with a sleep of provided seconds
        while "false" not in os.environ["perfecto_actions_refresh"]:
            time.sleep(int(os.environ["perfecto_actions_refresh"]))
            main()
        devlist.close()
        devlist.terminate()
        if(bool):
            if str(args["device_status"]).lower() in ["disconnected","notavailable","unavailable"]:
                print("There are some devices which are " + args["device_status"].lower())
                sys.exit(-1)
            
    except Exception as e:
        print(traceback.format_exc())
        sys.exit(-1)

def killProcess():
    try:
        if not platform.system() == "Darwin":
                os.system("taskkill /f /im perfectoactions.exe")
    except Exception as e:
        print("Killing process failed.." + str(e))
        
if __name__ == "__main__":
    try:
        main()
        print("Execution completed successfully!")
        sys.exit(0)
    except Exception as e:
        print(traceback.format_exc())
        sys.exit(-1)
    finally:
        killProcess()
